"""
Excel Comparator Module
A comprehensive module for comparing two Excel files and generating detailed difference reports.
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional
import warnings
warnings.filterwarnings('ignore')


class ExcelComparator:
    """Main class for comparing two Excel files."""

    def __init__(self, file1_path: str, file2_path: str, key_column: Optional[str] = None):
        """
        Initialize the comparator with two Excel files.

        Args:
            file1_path: Path to first Excel file
            file2_path: Path to second Excel file
            key_column: Optional column name to use as key for row alignment
        """
        self.file1_path = file1_path
        self.file2_path = file2_path
        self.key_column = key_column
        self.df1 = None
        self.df2 = None
        self.comparison_results = {}

    def load_files(self) -> Tuple[bool, str]:
        """
        Load both Excel files into pandas DataFrames.

        Returns:
            Tuple of (success: bool, message: str)
        """
        try:
            self.df1 = pd.read_excel(self.file1_path)
            self.df2 = pd.read_excel(self.file2_path)
            return True, "Files loaded successfully"
        except FileNotFoundError as e:
            return False, f"File not found: {str(e)}"
        except Exception as e:
            return False, f"Error loading files: {str(e)}"

    def align_by_key(self) -> None:
        """Align and sort both DataFrames by the key column if specified."""
        if self.key_column and self.key_column in self.df1.columns and self.key_column in self.df2.columns:
            self.df1 = self.df1.sort_values(by=self.key_column).reset_index(drop=True)
            self.df2 = self.df2.sort_values(by=self.key_column).reset_index(drop=True)

    def compare_structure(self) -> Dict:
        """
        Compare the structure of both DataFrames.

        Returns:
            Dictionary with structure comparison results
        """
        results = {
            'rows_match': len(self.df1) == len(self.df2),
            'rows_file1': len(self.df1),
            'rows_file2': len(self.df2),
            'cols_match': len(self.df1.columns) == len(self.df2.columns),
            'cols_file1': len(self.df1.columns),
            'cols_file2': len(self.df2.columns),
            'column_names_match': list(self.df1.columns) == list(self.df2.columns),
            'columns_file1': list(self.df1.columns),
            'columns_file2': list(self.df2.columns),
            'missing_in_file1': list(set(self.df2.columns) - set(self.df1.columns)),
            'missing_in_file2': list(set(self.df1.columns) - set(self.df2.columns)),
        }
        return results

    def compare_data(self) -> Tuple[pd.DataFrame, int]:
        """
        Compare data cell by cell and track all differences.

        Returns:
            Tuple of (differences_df: DataFrame, difference_count: int)
        """
        differences = []

        # Get common columns
        common_columns = list(set(self.df1.columns) & set(self.df2.columns))

        if not common_columns:
            return pd.DataFrame(), 0

        # Determine the number of rows to compare
        max_rows = max(len(self.df1), len(self.df2))

        for row_idx in range(max_rows):
            row_differences = {}
            has_difference = False

            # Add key column if exists
            if self.key_column and self.key_column in common_columns:
                key_val_1 = self.df1.at[row_idx, self.key_column] if row_idx < len(self.df1) else "MISSING"
                key_val_2 = self.df2.at[row_idx, self.key_column] if row_idx < len(self.df2) else "MISSING"
                row_differences[self.key_column] = key_val_1 if key_val_1 == key_val_2 else f"{key_val_1} / {key_val_2}"

            for col in common_columns:
                if col == self.key_column:
                    continue

                # Get values from both files
                val1 = self.df1.at[row_idx, col] if row_idx < len(self.df1) else np.nan
                val2 = self.df2.at[row_idx, col] if row_idx < len(self.df2) else np.nan

                # Compare values (handle NaN properly)
                if pd.isna(val1) and pd.isna(val2):
                    continue
                elif pd.isna(val1) or pd.isna(val2) or val1 != val2:
                    has_difference = True
                    row_differences[f'{col}_file1'] = val1
                    row_differences[f'{col}_file2'] = val2
                    row_differences[f'{col}_match'] = 'DIFF'

            # Only add rows with differences
            if has_difference:
                row_differences['row_index'] = row_idx
                differences.append(row_differences)

        differences_df = pd.DataFrame(differences)
        return differences_df, len(differences)

    def compare(self) -> Dict:
        """
        Perform complete comparison of both files.

        Returns:
            Dictionary containing all comparison results
        """
        # Load files
        success, message = self.load_files()
        if not success:
            return {'error': message}

        # Align by key if specified
        self.align_by_key()

        # Compare structure
        structure_results = self.compare_structure()

        # Compare data
        differences_df, diff_count = self.compare_data()

        # Determine if files are identical
        files_identical = (
            structure_results['rows_match'] and
            structure_results['cols_match'] and
            structure_results['column_names_match'] and
            diff_count == 0
        )

        results = {
            'identical': files_identical,
            'structure': structure_results,
            'differences_df': differences_df,
            'difference_count': diff_count,
        }

        self.comparison_results = results
        return results

    def generate_report(self, output_path: str = 'differences.xlsx') -> str:
        """
        Generate an Excel report with all differences.

        Args:
            output_path: Path for the output Excel file

        Returns:
            Success message or error message
        """
        if not self.comparison_results:
            return "No comparison results available. Run compare() first."

        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Write summary sheet
                summary_data = {
                    'Metric': [
                        'Files Identical',
                        'Rows in File 1',
                        'Rows in File 2',
                        'Rows Match',
                        'Columns in File 1',
                        'Columns in File 2',
                        'Columns Match',
                        'Column Names Match',
                        'Total Differences',
                    ],
                    'Value': [
                        self.comparison_results['identical'],
                        self.comparison_results['structure']['rows_file1'],
                        self.comparison_results['structure']['rows_file2'],
                        self.comparison_results['structure']['rows_match'],
                        self.comparison_results['structure']['cols_file1'],
                        self.comparison_results['structure']['cols_file2'],
                        self.comparison_results['structure']['cols_match'],
                        self.comparison_results['structure']['column_names_match'],
                        self.comparison_results['difference_count'],
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)

                # Write column comparison sheet
                if self.comparison_results['structure']['missing_in_file1'] or \
                   self.comparison_results['structure']['missing_in_file2']:
                    col_comparison = {
                        'Columns in File 1': pd.Series(self.comparison_results['structure']['columns_file1']),
                        'Columns in File 2': pd.Series(self.comparison_results['structure']['columns_file2']),
                    }
                    col_df = pd.DataFrame(col_comparison)
                    col_df.to_excel(writer, sheet_name='Column Comparison', index=False)

                # Write differences sheet
                if not self.comparison_results['differences_df'].empty:
                    diff_df = self.comparison_results['differences_df']
                    diff_df.to_excel(writer, sheet_name='Differences', index=False)

                    # Apply conditional formatting for better visualization
                    workbook = writer.book
                    worksheet = writer.sheets['Differences']

                    from openpyxl.styles import PatternFill, Font

                    # Highlight DIFF cells
                    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    bold_font = Font(bold=True)

                    for row in range(2, len(diff_df) + 2):
                        for col in range(1, len(diff_df.columns) + 1):
                            cell = worksheet.cell(row=row, column=col)
                            if cell.value == 'DIFF':
                                cell.fill = red_fill
                                cell.font = bold_font

            return f"Report generated successfully: {output_path}"
        except Exception as e:
            return f"Error generating report: {str(e)}"

    def print_summary(self) -> None:
        """Print a human-readable summary of the comparison."""
        if not self.comparison_results:
            print("No comparison results available. Run compare() first.")
            return

        print("\n" + "="*60)
        print("EXCEL FILE COMPARISON REPORT")
        print("="*60)

        if self.comparison_results['identical']:
            print("\n✓ FILES ARE IDENTICAL")
            print("\n" + "="*60)
            return

        print("\n📊 STRUCTURE COMPARISON:")
        print("-" * 60)

        struct = self.comparison_results['structure']

        # Row comparison
        if struct['rows_match']:
            print(f"✓ Rows: {struct['rows_file1']} (both files)")
        else:
            print(f"✗ Rows: File1={struct['rows_file1']}, File2={struct['rows_file2']}")

        # Column comparison
        if struct['cols_match']:
            print(f"✓ Columns: {struct['cols_file1']} (both files)")
        else:
            print(f"✗ Columns: File1={struct['cols_file1']}, File2={struct['cols_file2']}")

        # Column names comparison
        if struct['column_names_match']:
            print("✓ Column names and order match")
        else:
            print("✗ Column names or order differ")
            if struct['missing_in_file1']:
                print(f"  Missing in File 1: {', '.join(struct['missing_in_file1'])}")
            if struct['missing_in_file2']:
                print(f"  Missing in File 2: {', '.join(struct['missing_in_file2'])}")

        # Data differences
        print("\n📝 DATA COMPARISON:")
        print("-" * 60)
        diff_count = self.comparison_results['difference_count']
        if diff_count == 0:
            print("✓ All data matches")
        else:
            print(f"✗ Found {diff_count} rows with differences")

        print("\n" + "="*60)


def compare_excel_files(file1: str, file2: str, key_column: Optional[str] = None,
                        output_file: str = 'differences.xlsx') -> Dict:
    """
    Convenience function to compare two Excel files.

    Args:
        file1: Path to first Excel file
        file2: Path to second Excel file
        key_column: Optional column name to use as key for alignment
        output_file: Path for output differences file

    Returns:
        Dictionary with comparison results
    """
    comparator = ExcelComparator(file1, file2, key_column)
    results = comparator.compare()

    if 'error' not in results:
        comparator.print_summary()
        message = comparator.generate_report(output_file)
        print(f"\n{message}")
    else:
        print(f"\nError: {results['error']}")

    return results
